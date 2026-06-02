from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def format_excel(input_file, output_file):
    wb = load_workbook(input_file) # Load workbook

    # Theme colors
    HEADER_COLOR = "1F4E78"      # Dark blue
    HEADER_TEXT = "FFFFFF"       # White
    ALT_ROW_COLOR = "D9EAF7"     # Light blue
    BORDER_COLOR = "B7B7B7"

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    for ws in wb.worksheets:

        # Format header row
        for cell in ws[1]:
            cell.fill = PatternFill(
                start_color=HEADER_COLOR,
                end_color=HEADER_COLOR,
                fill_type="solid"
            )
            cell.font = Font(
                color=HEADER_TEXT,
                bold=True,
                size=12
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            cell.border = thin_border

        # Format data rows
        for row_num, row in enumerate(
            ws.iter_rows(min_row=2),
            start=2
        ):
            for cell in row:

                cell.border = thin_border

                # Alternate row shading
                if row_num % 2 == 0:
                    cell.fill = PatternFill(
                        start_color=ALT_ROW_COLOR,
                        end_color=ALT_ROW_COLOR,
                        fill_type="solid"
                    )

                cell.alignment = Alignment(
                    vertical="center"
                )

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    value = str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = min(max_length + 3, 50)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    wb.save(output_file)
    print(f"Formatted workbook saved as: {output_file}")
  
if __name__ == "__main__":
    input_excel   = "raw_data.xlsx"
    output_excel = "formatted_data.xlsx"

    format_excel(input_excel, output_excel)
